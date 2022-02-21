# Priority out.xlsx open file error fix for "expected <class 'openpyxl.styles.fills.Fill'>"
# "\AppData\Local\Programs\Python\Python310\Lib\site-packages\openpyxl\reader\excel.py"
# Add "if not self.data_only:" line 281 and on line 282 "apply_stylesheet(self.archive, self.wb)" 
# Use data_only=True on load_workbook(file_path, read_only=True, data_only=True)

import os
import configparser
import datetime

from openpyxl import Workbook
from openpyxl import load_workbook
from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
from tkinter.ttk import *

working_folder = os.getcwd()
window = Tk()
progress = Progressbar(window, orient=HORIZONTAL,
                       length=250)
config_path = os.path.join(os.getcwd(), 'config.ini')
config = configparser.ConfigParser()

if not os.path.isfile(config_path):
    config['row_acomulators'] = {'PO_COL': '9',
                                 'SO_COL': '11',
                                 'PK_COL': '15'}
    with open(config_path, 'w') as configfile:
        config.write(configfile)

config.sections()
config.read(config_path)
PO_COL = int(config['row_acomulators']['PO_COL']
             ) if config['row_acomulators']['PO_COL'] != "" else False
SO_COL = int(config['row_acomulators']['SO_COL']
             ) if config['row_acomulators']['SO_COL'] != "" else False
PK_COL = int(config['row_acomulators']['PK_COL']
             ) if config['row_acomulators']['PK_COL'] != "" else False

print(f"Settings {PO_COL},{SO_COL},{PK_COL}")


def is_int(s):
    try:
        int(s)
        return True
    except ValueError:
        return False


def save_workbook(rows):
    workbook = Workbook()
    sheet = workbook.active
    now = datetime.datetime.now()
    time_stamp = now.strftime("%d_%m_%y__%H_%M_%S")
    file_name = "stock_"+time_stamp+".xlsx"

    progress_step = 100/len(rows)
    for row in rows:
        if(progress['value'] > 100):
            progress['value'] = 0
        progress['value'] += progress_step
        window.update_idletasks()
        if len(row) > 0:
            sheet.append(row)

    sheet.freeze_panes = 'B2'
    print(f"output rows {sheet.max_row}")
    file_path = os.path.join(working_folder, file_name)
    print(file_path)
    progress['value'] = 0
    workbook.save(filename=file_path)
    os.startfile(working_folder)


def open_workbook(file_path):
    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        sheet = workbook.active
        print(f"input rows {sheet.max_row}")
    except PermissionError as e:
        print(f"{e}")
        messagebox.showwarning(
            'Error', f"{e} \n File opened in other app?", icon='error')
    except Exception as e:
        print(f"{e}")
        messagebox.showwarning(
            'Error', f"{e} \n File is broken?", icon='error')
        return

    rows = []
    all_rows = sheet.iter_rows(max_col=16, values_only=True)
    po_acomulator = []
    so_acomulator = []
    paka_acomulator = []
    prev, current = [""],[""]
    for row in all_rows:
        try:
            prev, current = current, row
            if PO_COL and isinstance(row[PO_COL], str) and row[PO_COL] != "":
                po_acomulator.append(row[PO_COL])

            if SO_COL and isinstance(row[SO_COL], str) and row[SO_COL] != "":
                so_acomulator.append(row[SO_COL])

            if PK_COL and isinstance(row[PK_COL], str) and row[PK_COL] != "":
                paka_acomulator.append(row[PK_COL])

            if current[0] != prev[0]:
                row_str = {}
                i = 1
                for cell in prev:
                    if i == PO_COL+1:
                        row_str[PO_COL+1] = ','.join(po_acomulator)
                        po_acomulator.clear()
                    elif i == SO_COL+1:
                        row_str[SO_COL+1] = ','.join(so_acomulator)
                        so_acomulator.clear()
                    elif i == PK_COL+1:
                        row_str[PK_COL+1] = ','.join(paka_acomulator)
                        paka_acomulator.clear()
                    else:
                        row_str[i] = cell
                    i += 1
                rows.append(row_str)
        except Exception as e:
            print(f"Error read row: {e} {row}")
            continue
    save_workbook(rows)



def run_app():
    global working_folder
    file_path = ""
    try:
        file_path = filedialog.askopenfilename(
            initialdir=working_folder, title="Open Priority Excel export file", filetypes=[("Only new Excel Files", "*.xlsx")])
        working_folder = os.path.dirname(file_path)
        if file_path != "":
            open_workbook(file_path)
    except Exception as e:
        print(f"{e}")
        messagebox.showwarning(
            'Error', f"{e}", icon='error')



def gui():
    print("Gui start")
    window.title("Priority Filter Rows")
    if os.path.isfile('icon.ico'):
        window.iconbitmap('icon.ico')
    lbl = Label(window, text="Progress")
    btn = Button(window, text='Start', command=run_app)
    progress.grid(column=1, row=1, pady=20, padx=10)
    lbl.grid(column=0, row=1, padx=10)
    btn.grid(column=0, row=2, padx=20, pady=20, columnspan=2)
    window.mainloop()


gui()
